"""
Service de génération de rapports.
Génère :
- Rapport PDF (via reportlab) : rapport complet avec en-tête, statistiques, table des tâches sans doublon
- Rapport Excel (via openpyxl) : rapport structuré multi-onglets
"""

import io
from datetime import datetime
from typing import List, Optional

# ── PDF ──────────────────────────────────────────────────────────────────────
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, KeepTogether
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

# ── Excel ─────────────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _trunc(text: str, max_len: int = 120) -> str:
    text = str(text or "")
    return text if len(text) <= max_len else text[:max_len] + "…"


# ─── PDF Report ───────────────────────────────────────────────────────────────

def generate_pdf_report(
    username: str,
    bulk_results: list,
    stats: dict,
    column_mapping: Optional[dict] = None,
    filename_source: Optional[str] = None,
) -> bytes:
    """
    Génère un rapport PDF complet.

    bulk_results : liste de {title, description, matches: [...]}
    stats : dict de statistiques globales
    """
    if not REPORTLAB_OK:
        raise RuntimeError(
            "reportlab n'est pas installé. Lancez : pip install reportlab"
        )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.8 * cm,
        leftMargin=1.8 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    INDIGO = colors.HexColor("#4F46E5")
    INDIGO_LIGHT = colors.HexColor("#EEF2FF")
    RED = colors.HexColor("#EF4444")
    ORANGE = colors.HexColor("#F97316")
    YELLOW = colors.HexColor("#EAB308")
    GREEN = colors.HexColor("#22C55E")
    GRAY = colors.HexColor("#64748B")
    DARK = colors.HexColor("#0F172A")

    style_h1 = ParagraphStyle("H1", parent=styles["Heading1"],
        fontSize=20, textColor=INDIGO, spaceAfter=4, fontName="Helvetica-Bold")
    style_h2 = ParagraphStyle("H2", parent=styles["Heading2"],
        fontSize=13, textColor=DARK, spaceAfter=6, fontName="Helvetica-Bold")
    style_h3 = ParagraphStyle("H3", parent=styles["Heading3"],
        fontSize=10, textColor=INDIGO, spaceAfter=4, fontName="Helvetica-Bold")
    style_body = ParagraphStyle("Body", parent=styles["Normal"],
        fontSize=9, leading=14, textColor=DARK)
    style_small = ParagraphStyle("Small", parent=styles["Normal"],
        fontSize=8, leading=12, textColor=GRAY)
    style_center = ParagraphStyle("Center", parent=styles["Normal"],
        fontSize=9, alignment=TA_CENTER)

    story = []

    # ── En-tête ──
    now = datetime.now().strftime("%d/%m/%Y à %Hh%M")
    story.append(Paragraph("⬡ Task Similarity Analyzer", style_h1))
    story.append(Paragraph("Rapport d'analyse des similarités et doublons", style_h2))
    story.append(HRFlowable(width="100%", thickness=2, color=INDIGO))
    story.append(Spacer(1, 0.3 * cm))

    # Méta-infos
    meta_data = [
        ["Utilisateur", username],
        ["Date du rapport", now],
        ["Fichier source", filename_source or "—"],
        ["Tâches soumises", str(stats.get("total_tasks_submitted", len(bulk_results)))],
        ["Doublons détectés", str(stats.get("total_duplicates_found", 0))],
        ["Taux de duplication", f"{stats.get('duplicate_rate', 0)}%"],
    ]
    if column_mapping:
        meta_data.append(["Colonne titre", column_mapping.get("title_col", "—")])
        meta_data.append(["Colonne description", column_mapping.get("description_col", "—")])

    meta_table = Table(meta_data, colWidths=[4 * cm, 12 * cm])
    meta_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), INDIGO_LIGHT),
        ("TEXTCOLOR", (0, 0), (0, -1), INDIGO),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (1, 0), (1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 0.5 * cm))

    # ── Résumé statistique ──
    story.append(Paragraph("Résumé de l'analyse", style_h2))

    total = len(bulk_results)
    dup_count = sum(1 for r in bulk_results if any(m["level"] == "doublon" for m in r["matches"]))
    strong_count = sum(1 for r in bulk_results
                       if not any(m["level"] == "doublon" for m in r["matches"])
                       and any(m["level"] == "forte" for m in r["matches"]))
    moderate_count = sum(1 for r in bulk_results
                         if not any(m["level"] in ("doublon", "forte") for m in r["matches"])
                         and r["matches"])
    clean_count = sum(1 for r in bulk_results if not r["matches"])

    stat_data = [
        ["Catégorie", "Nombre", "Proportion", "Action recommandée"],
        [Paragraph("🔴 Doublons (≥90%)", style_body), str(dup_count),
         f"{round(dup_count/total*100,1) if total else 0}%", "Supprimer ou fusionner"],
        [Paragraph("🟠 Similarité forte (70-89%)", style_body), str(strong_count),
         f"{round(strong_count/total*100,1) if total else 0}%", "Vérifier manuellement"],
        [Paragraph("🟡 Similarité modérée (50-69%)", style_body), str(moderate_count),
         f"{round(moderate_count/total*100,1) if total else 0}%", "Examiner si utile"],
        [Paragraph("✅ Tâches uniques", style_body), str(clean_count),
         f"{round(clean_count/total*100,1) if total else 0}%", "Conserver"],
    ]

    stat_table = Table(stat_data, colWidths=[6 * cm, 2 * cm, 3 * cm, 5.5 * cm])
    stat_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), INDIGO),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("ALIGN", (1, 0), (2, -1), "CENTER"),
    ]))
    story.append(stat_table)
    story.append(Spacer(1, 0.6 * cm))

    # ── Section : Tâches SANS doublon (rapport propre) ──
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E2E8F0")))
    story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph(f"✅ Tâches uniques à conserver ({clean_count})", style_h2))
    story.append(Paragraph(
        "Ces tâches n'ont aucune similarité significative avec les autres. Elles peuvent être conservées en l'état.",
        style_small
    ))
    story.append(Spacer(1, 0.3 * cm))

    clean_tasks = [r for r in bulk_results if not r["matches"]]
    if clean_tasks:
        clean_data = [["#", "Titre", "Description (extrait)"]]
        for i, r in enumerate(clean_tasks, 1):
            clean_data.append([
                str(i),
                Paragraph(_trunc(r["title"], 60), style_body),
                Paragraph(_trunc(r["description"], 100), style_small),
            ])
        t = Table(clean_data, colWidths=[0.8 * cm, 5.5 * cm, 10.2 * cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), GREEN),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0FDF4")]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1FAE5")),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ]))
        story.append(t)
    else:
        story.append(Paragraph("Aucune tâche unique détectée.", style_small))

    story.append(Spacer(1, 0.6 * cm))

    # ── Section : Doublons détectés ──
    dup_tasks = [r for r in bulk_results if any(m["level"] == "doublon" for m in r["matches"])]
    if dup_tasks:
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E2E8F0")))
        story.append(Spacer(1, 0.2 * cm))
        story.append(Paragraph(f"🔴 Doublons détectés ({len(dup_tasks)})", style_h2))
        story.append(Paragraph(
            "Ces tâches ont un score de similarité ≥ 90% avec une tâche existante.",
            style_small
        ))
        story.append(Spacer(1, 0.3 * cm))

        for r in dup_tasks:
            top_match = max(r["matches"], key=lambda m: m["similarity_score"])
            pct = round(top_match["similarity_score"] * 100, 1)
            block = [
                Paragraph(f"<b>{_trunc(r['title'], 80)}</b>", style_body),
                Paragraph(f"<font color='#64748B' size='8'>{_trunc(r['description'], 120)}</font>", style_small),
                Spacer(1, 0.1 * cm),
                Paragraph(
                    f"<font color='#EF4444'>→ Similaire à : <b>{_trunc(top_match['title'], 70)}</b> "
                    f"({pct}% — {', '.join(top_match['common_keywords'][:5])})</font>",
                    style_small
                ),
                HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#FECACA")),
                Spacer(1, 0.15 * cm),
            ]
            story.extend(block)

    # ── Section : Similarités fortes ──
    strong_tasks = [r for r in bulk_results
                    if not any(m["level"] == "doublon" for m in r["matches"])
                    and any(m["level"] == "forte" for m in r["matches"])]
    if strong_tasks:
        story.append(Spacer(1, 0.3 * cm))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E2E8F0")))
        story.append(Spacer(1, 0.2 * cm))
        story.append(Paragraph(f"🟠 Similarités fortes ({len(strong_tasks)})", style_h2))
        story.append(Paragraph(
            "Ces tâches ont un score entre 70% et 89%. Une vérification manuelle est recommandée.",
            style_small
        ))
        story.append(Spacer(1, 0.3 * cm))

        for r in strong_tasks:
            top_match = max(r["matches"], key=lambda m: m["similarity_score"])
            pct = round(top_match["similarity_score"] * 100, 1)
            block = [
                Paragraph(f"<b>{_trunc(r['title'], 80)}</b>", style_body),
                Paragraph(
                    f"<font color='#F97316'>→ {_trunc(top_match['title'], 70)} ({pct}%)</font>",
                    style_small
                ),
                HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#FED7AA")),
                Spacer(1, 0.1 * cm),
            ]
            story.extend(block)

    # ── Pied de page ──
    story.append(Spacer(1, 1 * cm))
    story.append(HRFlowable(width="100%", thickness=1, color=INDIGO))
    story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph(
        f"Rapport généré le {now} par Task Similarity Analyzer v1.0 — Confidentiel",
        style_small
    ))

    doc.build(story)
    return buffer.getvalue()


# ─── Excel Report ─────────────────────────────────────────────────────────────

def generate_excel_report(
    username: str,
    bulk_results: list,
    stats: dict,
    column_mapping: Optional[dict] = None,
    filename_source: Optional[str] = None,
) -> bytes:
    """Génère un rapport Excel multi-onglets."""
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl n'est pas installé.")

    wb = openpyxl.Workbook()

    # Couleurs
    INDIGO = "4F46E5"
    GREEN = "22C55E"
    RED = "EF4444"
    ORANGE = "F97316"
    YELLOW = "EAB308"
    LIGHT_GRAY = "F8FAFC"
    WHITE = "FFFFFF"

    def header_style(cell, bg=INDIGO):
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def auto_width(ws, min_w=10, max_w=60):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, min_w), max_w)

    thin = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    # ── Onglet 1 : Résumé ──
    ws1 = wb.active
    ws1.title = "Résumé"
    ws1.append(["Task Similarity Analyzer — Rapport d'analyse"])
    ws1["A1"].font = Font(bold=True, size=14, color=INDIGO)
    ws1.append([])
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws1.append(["Utilisateur", username])
    ws1.append(["Date", now])
    ws1.append(["Fichier source", filename_source or "—"])
    if column_mapping:
        ws1.append(["Colonne titre", column_mapping.get("title_col", "—")])
        ws1.append(["Colonne description", column_mapping.get("description_col", "—")])
        ws1.append(["Méthode détection", column_mapping.get("method", "—")])
    ws1.append([])
    ws1.append(["Statistiques", ""])
    ws1.append(["Total tâches analysées", len(bulk_results)])
    total = len(bulk_results)
    dup = sum(1 for r in bulk_results if any(m["level"] == "doublon" for m in r["matches"]))
    strong = sum(1 for r in bulk_results
                 if not any(m["level"] == "doublon" for m in r["matches"])
                 and any(m["level"] == "forte" for m in r["matches"]))
    mod = sum(1 for r in bulk_results
              if not any(m["level"] in ("doublon", "forte") for m in r["matches"])
              and r["matches"])
    clean = sum(1 for r in bulk_results if not r["matches"])
    ws1.append(["🔴 Doublons (≥90%)", dup])
    ws1.append(["🟠 Similarité forte (70-89%)", strong])
    ws1.append(["🟡 Similarité modérée (50-69%)", mod])
    ws1.append(["✅ Tâches uniques", clean])
    ws1.append(["Taux de duplication", f"{round(dup/total*100,1) if total else 0}%"])
    auto_width(ws1)

    # ── Onglet 2 : Tâches uniques (sans doublon) ──
    ws2 = wb.create_sheet("✅ Tâches uniques")
    headers2 = ["#", "Titre", "Description", "Statut"]
    ws2.append(headers2)
    for i, h in enumerate(headers2, 1):
        header_style(ws2.cell(1, i), GREEN)
    clean_tasks = [r for r in bulk_results if not r["matches"]]
    for i, r in enumerate(clean_tasks, 1):
        ws2.append([i, r["title"], r["description"], "Unique"])
        for col in range(1, 5):
            ws2.cell(i + 1, col).border = thin
            if i % 2 == 0:
                ws2.cell(i + 1, col).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["C"].width = 60
    ws2.column_dimensions["D"].width = 12
    ws2.row_dimensions[1].height = 20

    # ── Onglet 3 : Doublons ──
    ws3 = wb.create_sheet("🔴 Doublons")
    headers3 = ["#", "Titre soumis", "Description soumise", "Tâche similaire", "Score %", "Mots communs"]
    ws3.append(headers3)
    for i, h in enumerate(headers3, 1):
        header_style(ws3.cell(1, i), RED)
    dup_tasks = [r for r in bulk_results if any(m["level"] == "doublon" for m in r["matches"])]
    for i, r in enumerate(dup_tasks, 1):
        top = max(r["matches"], key=lambda m: m["similarity_score"])
        pct = round(top["similarity_score"] * 100, 1)
        ws3.append([i, r["title"], r["description"], top["title"], f"{pct}%", ", ".join(top["common_keywords"][:8])])
        for col in range(1, 7):
            ws3.cell(i + 1, col).border = thin
    ws3.column_dimensions["B"].width = 35
    ws3.column_dimensions["C"].width = 50
    ws3.column_dimensions["D"].width = 35
    ws3.column_dimensions["E"].width = 10
    ws3.column_dimensions["F"].width = 40

    # ── Onglet 4 : Toutes les similarités ──
    ws4 = wb.create_sheet("📊 Toutes similarités")
    headers4 = ["#", "Titre soumis", "Titre similaire", "Score %", "Niveau", "Mots-clés communs"]
    ws4.append(headers4)
    for i, h in enumerate(headers4, 1):
        header_style(ws4.cell(1, i))
    row_n = 2
    for i, r in enumerate(bulk_results, 1):
        if not r["matches"]:
            ws4.cell(row_n, 1).value = i
            ws4.cell(row_n, 2).value = r["title"]
            ws4.cell(row_n, 3).value = "—"
            ws4.cell(row_n, 4).value = "0%"
            ws4.cell(row_n, 5).value = "unique"
            ws4.cell(row_n, 6).value = ""
            ws4.cell(row_n, 5).fill = PatternFill("solid", fgColor="D1FAE5")
            row_n += 1
        else:
            for m in r["matches"]:
                pct = round(m["similarity_score"] * 100, 1)
                ws4.cell(row_n, 1).value = i
                ws4.cell(row_n, 2).value = r["title"]
                ws4.cell(row_n, 3).value = m["title"]
                ws4.cell(row_n, 4).value = f"{pct}%"
                ws4.cell(row_n, 5).value = m["level"]
                ws4.cell(row_n, 6).value = ", ".join(m["common_keywords"][:8])
                color = RED if m["level"] == "doublon" else (ORANGE if m["level"] == "forte" else "FDE68A")
                ws4.cell(row_n, 5).fill = PatternFill("solid", fgColor=color)
                ws4.cell(row_n, 5).font = Font(color=WHITE, bold=True)
                row_n += 1
    ws4.column_dimensions["B"].width = 35
    ws4.column_dimensions["C"].width = 35
    ws4.column_dimensions["D"].width = 10
    ws4.column_dimensions["E"].width = 12
    ws4.column_dimensions["F"].width = 45

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
